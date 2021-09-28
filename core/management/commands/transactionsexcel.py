from django.core.management.base import BaseCommand
from core.models.src.transaction import Transaction
from core.service.utils import toJalaliDateTime
from openpyxl import Workbook


class Command(BaseCommand):
    help = 'Get checkout Report Excel output'

    def handle(self, *args, **kwargs):
        workbook = Workbook()
        sheet = workbook.active
        headers = ["آیدی کیف پول", "آیدی تراکنش", "کد رهگیری", "مبلغ", "موجودی کیف پول در لحظه تراکنش", "وضعیت", "تاریخ"]

        for index, header in enumerate(headers, 1):
            sheet.cell(row=1, column=index, value=header)

        for index, transaction in enumerate(Transaction.objects(), 1):
            sheet.cell(row=index+1, column=1, value=transaction.walletId)
            sheet.cell(row=index+1, column=2, value=transaction.transactionId)
            sheet.cell(row=index+1, column=3, value=transaction.trackId)
            sheet.cell(row=index+1, column=4, value=transaction.amount)
            sheet.cell(row=index+1, column=5, value=transaction.walletCredit)
            sheet.cell(row=index+1, column=6, value=transaction.status)
            sheet.cell(row=index+1, column=7, value=toJalaliDateTime(transaction.createdAt))

        workbook.save('./transactions.xlsx')
