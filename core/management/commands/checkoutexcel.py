from django.core.management.base import BaseCommand
from core.models.src.checkout import CheckoutQueue
from core.service.utils import toJalaliDateTime
from openpyxl import Workbook


class Command(BaseCommand):
    help = 'Get checkout Report Excel output'

    def handle(self, *args, **kwargs):
        workbook = Workbook()
        sheet = workbook.active
        headers = ["آیدی تسویه", "آیدی کیف پول", "مبلغ", "شماره شبا", "نوع تسویه", "تاریخ"]
        for index, header in enumerate(headers, 1):
            sheet.cell(row=1, column=index, value=header)

        for index, checkout in enumerate(CheckoutQueue.objects(), 1):
            sheet.cell(row=index+1, column=1, value=checkout.checkoutId)
            sheet.cell(row=index+1, column=2, value=checkout.walletId)
            sheet.cell(row=index+1, column=3, value=checkout.amount)
            sheet.cell(row=index+1, column=4, value=checkout.iban)
            sheet.cell(row=index+1, column=5, value=checkout.checkoutDelay)
            sheet.cell(row=index+1, column=6, value=toJalaliDateTime(checkout.createdAt))

        workbook.save('./checkouts.xlsx')
